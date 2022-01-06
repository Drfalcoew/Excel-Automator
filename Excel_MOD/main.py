# Excel Organizer
# Drew Foster, Python, 11/21/2021
import re
from typing import Any  # idk what these are. i needed to import them to specify types Bool and Any for some reason..?
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.descriptors import Bool
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import pandas as pd

fedex_TP_worksheet: Any
fedex_PP_worksheet: Any
ups_worksheet: Any
fedex_kohls_worksheet: Any
ups_kohls_worksheet: Any

#  getting info about workbook and active worksheet
my_wb = load_workbook('_original.xlsx')
_original_ws = my_wb.active

main_wb: Workbook
main_ws: Any

end_row = _original_ws.max_row
end_col = _original_ws.max_column
pr_header = False
red_header = False

fedex_sheet_titles = ['FEDEX_TP', 'FEDEX_PP', 'RED', 'FEDEX_KOHLS']

#  initializing var to blue filler color
masterpack_fill = PatternFill(start_color='8497b0',
                              end_color='8497b0',
                              fill_type='solid')
split_fill = PatternFill(start_color='66FF66',
                         end_color='66FF66',
                         fill_type='solid')
bad_address_fill = PatternFill(start_color='C41E3A',
                               end_color='C41E3A',
                               fill_type='solid')


def remove_batched(_sorted):
    main_wb = load_workbook('_delete.xlsx')
    main_ws = main_wb.active
    _end_row = _original_ws.max_row
    temp = _end_row
    sheet = 'Sheet1'

    writer = pd.ExcelWriter('_delete_1.xlsx', engine='openpyxl')

    print("Starting...")
    print(_sorted)

    for row in range(2, _original_ws.max_row):
        batch_num = main_ws['AC' + str(row)].value
        print('row: {}, batch: {}'.format(row, batch_num))
        if batch_num == 0:
            continue
        min_del = row
        break
    try:
        print(f"MIN_DEL: {min_del}")
        _sorted.drop(index=_sorted.index[min_del - 2:_original_ws.max_row], axis=0, inplace=True)
        _sorted.to_excel(writer, 'Sheet1', index=False)
        writer.save()
        print(_sorted)

    except Exception as e:
        print(f"Cannot delete rows, error: {e}")
        print(f"Can't sort writer to excel: {e}")
        # filter_red()

    print("Finished")


def filter_red_new():
    _end_row = _original_ws.max_row
    temp = _end_row
    _file_name = '_original.xlsx'
    sheet = 'Sheet1'

    book = load_workbook(_file_name)
    df = pd.read_excel(_file_name, engine='openpyxl', sheet_name=sheet)
    writer = pd.ExcelWriter('_delete.xlsx', engine='openpyxl')
    writer.book = book

    print("Starting...")

    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    _sorted = df.sort_values('Batch #')

    min_del: int  # Use this to determine range for splits eg: for i in range(min_del): check_if_split(row[i]): splits.append(order_number_of_row)

    print(_original_ws.max_row)

    print(_sorted)

    _sorted.to_excel(writer, 'Sheet1', index=False)
    try:
        writer.save()
    except Exception as e:
        print(f"ERROR: Cannot save writer, line 115. Exception: {e}")
    else:
        remove_batched(_sorted)


def filter_red():
    red = "FFFF0000"
    black = "FF424649"
    _end_row = _original_ws.max_row
    temp = _end_row
    print("Starting...")
    for row in range(1, _end_row):
        if _original_ws['A' + str(temp)].font.color.rgb == red:
            _original_ws.delete_rows(temp)
            print('A{} font: red. Deleting'.format(temp))
        elif _original_ws['A' + str(temp)].font.color.rgb == black:
            print('A{} font: black. Continuing'.format(temp))
        temp -= 1
    print("...Finished")


def filter_hot():
    value_list = []
    header_row = []
    global red_header
    global main_wb

    print("FILTERING HOT ORDERS")
    temp = main_ws.max_row
    for row in range(1, main_ws.max_row):
        value_list.clear()
        spec = main_ws['AH' + str(temp)].value
        if spec == "SHIP VIA OVERNIGHT" or spec == "SHIP VIA 2ND  DAY":
            print('AH{} SPEC INSTR: {}. HOT ORDER'.format(temp, spec))

            if 'RED' not in main_wb.sheetnames:
                main_wb.create_sheet("RED")
                if not red_header:
                    for i in range(1, 43):
                        header_row.append(main_ws[str(get_column_letter(i)) + str(1)].value)
                    try:
                        main_wb["RED"].append(header_row)
                    except:
                        print("main_wb['RED'] could not be appended to. possibly doesn't exist?")
                    red_header = True

            print(f"Moving row over to sheet {main_wb['RED']}")
            for i in range(1, 43):  # Copying original cell values
                value_list.append(main_ws[str(get_column_letter(i)) + str(temp)].value)
            main_ws.delete_rows(temp)
            move_over_sheet(value_list, main_wb['RED'])
        temp -= 1


def move_over_sheet(data, sheet):
    sheet.append(data)


def filter_freight():
    global main_wb
    global main_ws

    main_wb = load_workbook("_delete_1.xlsx")
    main_ws = main_wb.active

    tp = "TP"
    p = "P"
    pi = "PI"
    _end_row = main_ws.max_row
    temp = _end_row
    print("Starting...")
    for row in range(1, _end_row):
        freight_value = main_ws['Y' + str(temp)].value
        if freight_value == "" or freight_value is None:
            print('A{} FREIGHT: {}. Continuing'.format(temp, freight_value))
        elif freight_value == tp or freight_value == p or freight_value == pi:
            print('A{} FREIGHT: {}. Continuing'.format(temp, freight_value))
        else:
            main_ws.delete_rows(temp)
            print('A{} FREIGHT: {}. Deleting'.format(temp, freight_value))
        temp -= 1
    print("...Finished")

    save()


def filter_customer_name():
    dunhams = "DUNHAMS SPORTS EDI"
    bj = "BJ'S WHOLESALE CLUB"
    walmart = "WALMART.COM EDI INVENTORY"
    scheels = "SCHEELS EDI"
    walmart_1 = "WAL - MART - --- EDI"
    aafes = "AAFES CATALOG SALES"
    qvc = "QVC EDI ONLY"

    _end_row = main_ws.max_row
    temp = _end_row
    print("Starting...")
    for row in range(1, _end_row):
        customer_name = main_ws['AP' + str(temp)].value
        if customer_name == aafes:
            main_ws['AL' + str(temp)] = "FXGR"
            print('AP{} Customer Name: {}. Marking as fedex'.format(temp, customer_name))
        elif customer_name == qvc:
            main_ws['X' + str(temp)].value = "UPS"
        elif customer_name == dunhams or customer_name == bj or customer_name == walmart or customer_name == scheels or customer_name == walmart_1:
            print('AP{} Customer Name: {}. Deleting'.format(temp, customer_name))
            main_ws.delete_rows(temp)
        else:
            print('A{} Customer Name: {}. Continuing'.format(temp, customer_name))
        temp -= 1
    print("...Finished")


def filter_spec():
    value_list = []
    pilot = "PILOT"
    esfww = "SHIP VIA EFWW-ESTES FORWARDING WW"
    ground = "GROUND"
    home = "HOME"

    kohls = "KOHLS DROP SHIP EDI"
    _end_row = main_ws.max_row
    temp = _end_row
    sheet: Any
    transfer: Bool
    print("Starting...")
    for row in range(1, _end_row):
        transfer = False
        value_list.clear()
        spec = main_ws['AH' + str(temp)].value
        cust = main_ws['AP' + str(temp)].value
        freight = main_ws['Y' + str(temp)].value
        if spec == "" or spec is None:
            print('AH{} SPEC INSTR: {}. Continuing'.format(temp, spec))
        elif spec == esfww or pilot in spec:
            print('AH{} SPEC INSTR: {}. Deleting'.format(temp, spec))
            main_ws.delete_rows(temp)
        else:
            if "FED" in spec:  # check if special instructions contains the word "FED"
                transfer = True
                if ground in spec or "GND" in spec:
                    main_ws['AH' + str(temp)] = "FEDEX GROUND"
                elif home in spec:
                    main_ws['AH' + str(temp)] = "FEDEX HOME DELIVERY"

                if cust == kohls:
                    sheet = fedex_kohls_worksheet
                else:
                    if freight == 'TP':
                        sheet = fedex_TP_worksheet
                    else:
                        sheet = fedex_PP_worksheet

            elif spec == "SHIP VIA UPS GROUND":
                transfer = True
                if cust == kohls:
                    sheet = ups_kohls_worksheet
                else:
                    sheet = ups_worksheet

        if transfer:
            print(f"Moving row over to sheet {sheet}")
            for i in range(1, 43):  # Copying original cell values
                value_list.append(main_ws[str(get_column_letter(i)) + str(temp)].value)
            main_ws.delete_rows(temp)
            move_over_sheet(value_list, sheet)
            print('AH{} SPEC INSTR: {}. Continuing'.format(temp, spec))
        temp -= 1
    print("...Finished")


def filter_zero_weights():
    global main_wb
    global main_ws

    _end_row = main_ws.max_row
    temp = _end_row
    print("Starting...")
    for row in range(1, _end_row):
        weight = main_ws['G' + str(temp)].value
        _master_p = main_ws['H{0}'.format(str(temp))].value
        if weight == 0 or weight == 0.0 or weight == .0 or weight is None or weight == 0.01:
            print('G{} Weight: {}. Deleting'.format(temp, weight))
            main_ws.delete_rows(temp)
        elif 0.1 <= weight < 1:  # Between 0 and 1
            main_ws['G' + str(temp)] = 1
        else:  # Over 1 lb
            print('G{} Weight: {}. Continuing'.format(temp, weight))
            main_ws['G' + str(temp)] = int("{:.0f}".format(weight))  # Formatting weight to 0 decimals

        if _master_p == 0:
            main_ws['H{0}'.format(str(temp))] = 1
        temp -= 1
    print("...Finished")


def filter_char_osadx():
    global main_ws
    _end_row = main_ws.max_row
    temp = _end_row
    print("Starting... (removing non-numerics from phone #s)")
    for row in range(1, _end_row):
        osadx = main_ws['M' + str(temp)].value
        if osadx != "" or osadx is not None:
            main_ws['M' + str(temp)] = re.sub("[^0-9]", "", str(osadx))
            if len(str(osadx)) != 10:
                main_ws['M' + str(temp)] = ""
        temp -= 1
    print("...Finished")


def get_data_coord() -> tuple:
    start = 'A1'
    _end_col = get_column_letter(main_ws.max_column)
    _end_row = main_ws.max_row
    end = (_end_col + str(_end_row))
    return start, end


def filter_ship_via():
    value_list = []
    fx = "FEDEX"
    ups = "UP"
    pilot = "PILOT"
    ceva = "CEVA"
    _end_row = main_ws.max_row
    temp = _end_row
    sheet: Any
    transfer: Bool
    print("Starting... (SHIP VIA)")
    for row in range(1, _end_row):
        transfer = False
        sheet = None
        ship_via_value = main_ws['X' + str(temp)].value
        freight = main_ws['Y' + str(temp)].value

        if ship_via_value == "" or ship_via_value is None:
            print('A{} SCAC: {}. Continuing'.format(temp, ship_via_value))
        elif ship_via_value == fx:
            #  move to fedex sheet
            transfer = True
            if freight == 'TP':
                sheet = fedex_TP_worksheet
            else:
                sheet = fedex_PP_worksheet
            print('A{} SHIP_VIA: {}. Transferring data'.format(temp, ship_via_value))
        elif ship_via_value[0:2] == ups:
            #  move to ups sheet
            transfer = True
            sheet = ups_worksheet
            print('A{} SHIP_VIA: {}. Transferring data'.format(temp, ship_via_value))
        elif ship_via_value == pilot or ship_via_value == ceva:
            main_ws.delete_rows(temp)

        if transfer:
            value_list.clear()
            for i in range(1, 43):
                value_list.append(main_ws[str(get_column_letter(i)) + str(temp)].value)
            print(value_list)
            main_ws.delete_rows(temp)
            move_over_sheet(value_list, sheet)

        temp -= 1
    print("...Finished")


def filter_scac():
    #  move fedex and UPS to separate sheets
    value_list = []
    fedex = "FXGR"
    ups = "UP"
    _end_row = main_ws.max_row
    temp = _end_row
    sheet: Any
    transfer: Bool
    ground = "GROUND"
    home = "HOME"

    print("Starting...")
    for row in range(1, _end_row):
        transfer = False
        sheet = None
        scac_value = main_ws['AL' + str(temp)].value
        spec = main_ws['AH' + str(temp)].value
        freight = main_ws['Y' + str(temp)].value
        if scac_value == "" or scac_value is None:
            print('A{} SCAC: {}. Continuing'.format(temp, scac_value))
        elif scac_value == fedex:
            #  move to fedex sheet
            transfer = True
            if spec is None or spec is '':
                pass
            else:
                print(f"SPEC : {spec}")
                if "FED" in spec:  # check if special instructions contains the word "FED"
                    transfer = True
                    if ground in spec or "GND" in spec:
                        main_ws['AH' + str(temp)] = "FEDEX GROUND"
                    elif home in spec:
                        main_ws['AH' + str(temp)] = "FEDEX HOME DELIVERY"

            if freight == 'TP':
                sheet = fedex_TP_worksheet
            else:
                main_ws['AJ' + str(temp)] = '396428580'
                main_ws['AK' + str(temp)] = '570'
                sheet = fedex_PP_worksheet
            print('A{} SCAC: {}. Transferring data'.format(temp, scac_value))
        elif scac_value[0:2] == ups:
            #  move to ups sheet
            transfer = True
            sheet = ups_worksheet
            print('A{} SCAC: {}. Continuing'.format(temp, scac_value))
        else:
            main_ws.delete_rows(temp)
            print('A{} SCAC: {}. Deleting'.format(temp, scac_value))

        if transfer:
            value_list.clear()
            for i in range(1, 43):
                value_list.append(main_ws[str(get_column_letter(i)) + str(temp)].value)
            print(value_list)
            main_ws.delete_rows(temp)
            move_over_sheet(value_list, sheet)

        temp -= 1
    print("...Finished")


def duplicate_cartons(sheet):
    value_list = []
    _end_row = sheet.max_row
    temp = _end_row - 1
    print(f"Starting... (duplicating 2x) on sheet {sheet}")
    for row in range(1, _end_row - 1):
        open_q = sheet['F' + str(temp)].value
        master_p = sheet['H{0}'.format(str(temp))].value
        if open_q > master_p == 1:
            print(f"Row {temp} duplicating")
            r = int(open_q / master_p) - 1
            print(r)
            sheet['F' + str(temp)] = 1
            for dup in range(r):
                value_list.clear()
                for i in range(1, 43):
                    value_list.append(sheet[str(get_column_letter(i)) + str(temp)].value)
                print(value_list)
                sheet.append(value_list)
            print(value_list)
        temp -= 1
    print("...Finished")
    save()


def masterpacks(sheet):
    _end_row = sheet.max_row
    temp = _end_row - 1
    value_list = []

    print(f"Starting... (Checking for MPs) on sheet {sheet}")
    for _ in range(1, _end_row - 1):
        open_q = sheet['F' + str(temp)].value
        prod = sheet['E' + str(temp)].value

        master_p = sheet['H{0}'.format(str(temp))].value

        if open_q is not None and open_q > 1 and master_p != 1:
            print(f"Row {temp} is a masterpack")
            # sheet['G' + str(temp)].fill = masterpack_fill  # filling OG cell range with blue color

            if open_q % master_p == 0 or open_q > master_p:
                sheet['E' + str(temp)] = f"{prod}XMP"
            else:  # quantity is less than masterpack, but still > 1
                sheet['E' + str(temp)] = f"{prod}X{open_q}"

            if open_q > master_p:  # Need to duplicate cell
                # change original cell
                sheet['F' + str(temp)] = sheet['H' + str(temp)].value
                # need to duplicate cells.
                r: int = int(open_q / master_p)

                if open_q % master_p == 0:  # Complete masterpack
                    for dup in range(r - 1):
                        value_list.clear()
                        for i in range(1, 43):  # Copying original cell values
                            value_list.append(sheet[str(get_column_letter(i)) + str(temp)].value)
                        # Updating duplicated cell values
                        value_list[5] = master_p
                        value_list[4] = f"{prod}XMP"

                        print(value_list)
                        sheet.append(value_list)
                        # sheet['G' + str(sheet.max_row)].fill = masterpack_fill  # filling OG cell range with blue color

                else:  # Broken masterpack
                    for dup in range(r):
                        value_list.clear()
                        for i in range(1, 43):
                            value_list.append(sheet[str(get_column_letter(i)) + str(temp)].value)
                        if dup == r - 1:  # if increment is on final loop
                            value_list[5] = open_q % master_p
                            value_list[4] = f"{prod}X{value_list[5]}"

                        else:
                            value_list[5] = master_p
                            value_list[4] = f"{prod}XMP"

                        print(value_list)
                        sheet.append(value_list)
        temp -= 1
    print("...Finished")
    save()


def color_masterpacks(_my_wb):
    print("... Starting to color masterpacks blue")

    for sheet in _my_wb.worksheets:

        if sheet.title == 'Sheet1':
            continue

        _end_row = sheet.max_row
        temp = _end_row

        for _ in range(1, _end_row):
            open_q = sheet['F' + str(temp)].value
            weight = sheet['G' + str(temp)].value
            if open_q == None or open_q == "":
                continue

            if open_q > 1:
                print(f"Row {temp} is being filled on sheet: {sheet}")
                sheet['G' + str(temp)].fill = masterpack_fill  # filling cell with color
                sheet['G' + str(temp)] = weight * sheet['F' + str(temp)].value  # multiplying weight * quantity

            temp -= 1
    print("Finished...")


def duplicate_masterpack():
    for sheet in main_wb.worksheets:
        if sheet.title == 'Sheet1':
            continue
        duplicate_cartons(sheet)
        masterpacks(sheet)


def move_headers():
    global main_wb
    header_row = []
    header_width = []
    # Get header data
    for i in range(1, 43):
        header_row.append(_original_ws[str(get_column_letter(i)) + str(1)].value)
        header_width.append(_original_ws.column_dimensions[get_column_letter(i)].width)

    for sheet in main_wb.worksheets:
        sheet.append(header_row)
        for i in range(1, 43):
            sheet.column_dimensions[get_column_letter(i)].width = header_width[i - 1]


def save():
    try:
        main_wb.save('_delete_1.xlsx')
    except:
        print("ERROR saving to '_delete_1.xlsx'")


def sort_sheets():  # this screws up blue color for masterpacks, need to fill color only after sorts are complete
    print("SORTING BULLSHIT")
    _file_name = '_delete_1.xlsx'

    for sheet in main_wb.worksheets:
        book = load_workbook(_file_name)
        df = pd.read_excel(_file_name, engine='openpyxl', sheet_name=sheet.title)
        writer = pd.ExcelWriter(_file_name, engine='openpyxl')
        writer.book = book
        sorted : Any

        ## ExcelWriter for some reason uses writer.sheets to access the sheet.
        ## If you leave it empty it will not know that sheet Main is already there
        ## and will create a new sheet.

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        if sheet.title == 'UPS':
            sorted = df.sort_values(['CUSNO', 'PRDNO'])
        else:
            sorted = df.sort_values('PRDNO')
        print(sorted)
        sorted.to_excel(writer, sheet.title, index=False)

        writer.save()

        print("FINISHED SORTING")


def create_batches():
    print("-----CREATING BATCHES------")
    global main_wb
    global fedex_PP_worksheet
    worksheet = fedex_PP_worksheet  # assigning here just for testing. Later I will need to loop through each sheet, pref from outside the func sending a parameter.. cleaner
    value_list = []

    count: int  # Going to use this to count how many in a batch. if prod == nxt_prod * 10, insert line, otherwise... --> ?
    prod: str  # product name "PRDNO"
    nxt_prod: str

    _end_row = worksheet.max_row
    temp = _end_row  # I think I can increment from top down on this one. otherwise (down to up) use temp to control flow
    # I need to make sure to insert line on top of product * 10. so insert(row - count) (top) as well as insert(row) (bottom)

    # you could do..., first get all prod != nxtProd row lines, then in another loop, do insert row at line_breaks[i] + i (5 + 0, 24 + 1, 65 + 2)
    for row in range(2, _end_row):
        temp -= 1
    if worksheet['E' + str(temp)].value != worksheet['E' + str((temp - 1))].value:
        print(temp, worksheet['E' + str(temp)].value, worksheet['E' + str((temp - 1))].value)
        print(f"Inserting row at {temp}")
        worksheet.insert_rows(temp)

    save()


def gather_split_orders():
    split_orders = list()
    green = "FF66FF66"
    print("GATHERING SPLIT ORDERS __________")

    sheet = my_wb.active
    _end_row = sheet.max_row
    temp = _end_row

    for _ in range(1, _end_row - 1):
        cell_bg = sheet['A' + str(temp)].fill.start_color.index
        print(cell_bg)
        if cell_bg == green:
            print(f"Row {temp} is being collected on sheet: {sheet}")
            split_orders.append(sheet['B' + str(temp)].value)

        temp -= 1

    return set(split_orders)


def color_splits(_my_wb, split_orders):
    if split_orders is not None:
        print(split_orders)
        for sheet in _my_wb.worksheets:
            _end_row = sheet.max_row

            if sheet.title == 'Sheet1':
                continue
            for row in range(2, _end_row + 1):
                _order = sheet['B' + str(row)].value
                if _order in split_orders:
                    sheet['A' + str(row)].fill = split_fill


def filter_PR():
    global pr_header
    global main_wb
    global main_ws

    header_row = []
    value_list = []
    _end_row = main_ws.max_row
    temp = _end_row
    print("Starting...")
    for row in range(1, _end_row):
        state = main_ws['O' + str(temp)].value
        if state == "PR":
            print('O{} State: {}. Moving'.format(temp, state))
            value_list.clear()
            if "PR" not in main_wb.sheetnames:
                main_wb.create_sheet("PR")

                if not pr_header:
                    for i in range(1, 43):
                        header_row.append(main_ws[str(get_column_letter(i)) + str(1)].value)
                    try:
                        main_wb["PR"].append(header_row)
                    except:
                        print("main_wb['PR'] could not be appended to. possibly doesn't exist?")
                    pr_header = True

            print(f"Moving row over to sheet {main_wb['PR']}")
            for i in range(1, 43):  # Copying original cell values
                value_list.append(main_ws[str(get_column_letter(i)) + str(temp)].value)
            main_ws.delete_rows(temp)
            try:
                move_over_sheet(value_list, main_wb['PR'])
            except:
                print("Failed to move row over to 'PR'")
        temp -= 1
    print("...Finished")


def rewrite_ups(_my_wb):
    #  GROUND THIRD PARTY PROFILE PACKAGE PACKAGE
    sheets = ['UPS', 'UPS_KOHLS']

    for sheet in _my_wb.worksheets:
        if sheet.title not in sheets:
            continue  # UPS and UPS_KOHLS worksheets only.
        _end_row = sheet.max_row
        temp = _end_row
        print("Starting...")
        for row in range(1, _end_row):
            freight = sheet['Y' + str(temp)].value

            if freight == 'TP':
                sheet['Y' + str(temp)] = "THIRD PARTY"
            elif 'P' in freight:
                sheet['Y' + str(temp)] = "SHIPPER"

            sheet['X' + str(temp)].value = 'GROUND'
            sheet['Z' + str(temp)].value = ''
            sheet['AA' + str(temp)].value = 'PACKAGE'
            sheet['AB' + str(temp)].value = 'PACKAGE'

            temp -= 1
        print("...Finished")


def create_sheets():
    #  setting up new worksheets for fedex and UPS
    global main_wb
    global fedex_TP_worksheet
    global fedex_PP_worksheet
    global ups_worksheet
    global fedex_kohls_worksheet
    global ups_kohls_worksheet

    main_wb.create_sheet("FEDEX_TP")
    main_wb.create_sheet("FEDEX_PP")
    main_wb.create_sheet("UPS")
    main_wb.create_sheet("FEDEX_KOHLS")
    main_wb.create_sheet("UPS_KOHLS")

    fedex_TP_worksheet = main_wb["FEDEX_TP"]
    fedex_PP_worksheet = main_wb["FEDEX_PP"]
    ups_worksheet = main_wb["UPS"]
    fedex_kohls_worksheet = main_wb["FEDEX_KOHLS"]
    ups_kohls_worksheet = main_wb["UPS_KOHLS"]


def rewrite_fedex(_my_wb):
    for sheet in _my_wb.worksheets:
        if "FED" not in sheet.title:
            continue
        _end_row = sheet.max_row
        sheet.move_range(f"AJ1:AK{end_row}", rows=0, cols=-12)
        sheet.delete_cols(26, 7)
        sheet.delete_cols(31, 6)


def color_and_rewrite(split_orders):
    _my_wb = load_workbook("_delete_1.xlsx")

    format_spec(_my_wb)
    check_bad_address(_my_wb)
    color_masterpacks(_my_wb)
    color_splits(_my_wb, split_orders)
    rewrite_ups(_my_wb)
    rewrite_fedex(_my_wb)

    try:
        _my_wb.save('_sorted.xlsx')
    except:
        print("ERROR saving to '_sorted.xlsx'")

def format_spec(_my_wb):
    for sheet in _my_wb.worksheets:
        if sheet.title in fedex_sheet_titles:
            max_row = sheet.max_row
            for row in range(1, max_row+1):
                print(f"FORMATTING SPEC IN SHEET: {sheet.title} on ROW: {row}")

                if sheet['AH' + str(row)].value == "" or sheet['AH' + str(row)].value is None:
                    sheet['AH' + str(row)] = "FEDEX GROUND"


def check_bad_address(my_wb):
    global main_wb
    for sheet in my_wb.worksheets:

        _end_row = sheet.max_row
        temp = _end_row
        print("Starting... ")
        for row in range(1, _end_row):
            osad1 = sheet['K' + str(temp)].value

            if osad1 is not None and osad1[0].isdigit() == False:
                print("(Checking bad addresses. Filling with red)")
                sheet['K' + str(temp)].fill = bad_address_fill
            temp -= 1
        print("...Finished")


def cleanup():
    try:
        os.remove("_delete.xlsx")
        os.remove("_delete_1.xlsx")
    except:
        print("File removed successfully")


def filter_main():
    split_orders = gather_split_orders()  # fkn pandas removes colors so I have to check for splits before I clean up the spreadsheet, slowing it down.  (actually it's pretty fast)
    # BUT I can check for splits legit after cleaning up by checking on each sheet, nested for loop, outer nest get orderA, inner loop check if orderA = orderB-Z
    filter_red_new()

    # filter_red()  # slow as hell, deprecated

    filter_freight()  # TP, PI, P
    filter_zero_weights()
    filter_customer_name()
    filter_char_osadx()

    create_sheets()
    move_headers()

    # Now transferring or deleting
    filter_hot()
    filter_PR()
    filter_scac()  # FXGR, UPS, etc.
    filter_spec()  # Special Instructions
    filter_ship_via()

    duplicate_masterpack()  # Duplicating cartons / masterpacks on each sheet

    save()  # save _delete.xlsx

    sort_sheets()
    #create_batches()  this causes sort_sheets() not to work. ? (I think)

    color_and_rewrite(split_orders)  # Saves to new sheet called '_sorted.xlsx'

    cleanup()


filter_main()
